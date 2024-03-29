from datetime import timedelta
from logging import getLogger

import openpyxl
from constance import config
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.urls import reverse
from django.utils.html import format_html
from django.utils.timezone import now
from django.views.decorators.http import require_http_methods
from django.views.generic import CreateView, ListView, TemplateView, UpdateView

from accounts.models import User
from entries.forms import EntryForm
from entries.helpers import process_action
from entries.models import Entry, Rule

logger = getLogger(__name__)

ACCEPTABLE_FILE_EXTENSIONS = [
    '.xlsx'
]


COLOR_PALETTE = [
    '#FFEC21',
    '#378AFF',
    '#FFA32F',
    '#F54F52',
    '#93F03B',
    '#9552EA',
    '#ffffff',
    '#000000',
    '#aaaaaa',
    '#cccccc',
    '#666666',
]


class LogView(UserPassesTestMixin, ListView):
    template_name = 'entries/log.html'
    model = Entry
    context_object_name = 'entries'
    queryset = Entry.objects.all()
    paginate_by = 25

    def test_func(self):
        if config.PUBLIC_MODLOG:
            return True
        elif self.request.user.is_authenticated:
            return True
        return False


class RulesView(ListView):
    template_name = 'entries/rules.html'
    model = Rule
    context_object_name = 'rules'
    queryset = Rule.objects.all()


class Search(UserPassesTestMixin, ListView):
    template_name = 'entries/search.html'
    model = Entry
    context_object_name = 'entries'
    paginate_by = 25

    def test_func(self):
        if config.PUBLIC_MODLOG:
            return True
        elif self.request.user.is_authenticated:
            return True
        return False

    def get_queryset(self):
        query = self.request.GET.get('q')
        if not query:
            return Entry.objects.none()
        return Entry.objects.filter(
            Q(user__icontains=query) | Q(notes__icontains=query)
        )


class UserView(UserPassesTestMixin, ListView):
    template_name = 'entries/user.html'
    model = Entry
    context_object_name = 'entries'
    paginate_by = 25

    def test_func(self):
        if config.PUBLIC_MODLOG or self.request.user.is_authenticated:
            return True
        return False

    def get_queryset(self):
        username = self.kwargs.get('username')
        return Entry.objects.filter(
            Q(user__icontains=username) | Q(notes__icontains=username)
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['username'] = self.kwargs.get('username')
        return context


class AddEntryView(LoginRequiredMixin, CreateView):
    template_name = 'entries/add_entry.html'
    model = Entry
    form_class = EntryForm

    def form_valid(self, form):
        form.instance.moderator = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('log-view')


class EditEntryView(UserPassesTestMixin, UpdateView):
    template_name = 'entries/edit_entry.html'
    model = Entry
    form_class = EntryForm

    def get_success_url(self):
        return reverse('log-view')

    def test_func(self):
        """Allow only superusers and the adding moderator to edit."""
        obj = self.get_object()  # Need to do this because it is run before self.object is assigned.
        if self.request.user.is_superuser or self.request.user == obj.moderator:
            return True
        else:
            return False


class ImportEntriesView(LoginRequiredMixin, TemplateView):
    template_name = 'entries/import.html'

    def post(self, request, *args, **kwargs):
        message = ''
        file_object = request.FILES.get('file', None)
        success = True

        if f'.{file_object.name.split(".")[-1]}' not in ACCEPTABLE_FILE_EXTENSIONS:
            logger.error(f'User attempted to upload {file_object.name}.')
            success = False
            valid_extensions = ', '.join(ACCEPTABLE_FILE_EXTENSIONS)
            message = f'{file_object.name} is not a valid upload type. Please return to the previous page and try ' \
                      f'your upload again using one of the valid file extensions: {valid_extensions}'
        else:
            if Entry.objects.count() > 0:
                success = False
                message = 'The database has existing log entries. The import function can only be run when no ' \
                          'entries have been previously added.'
            else:
                workbook = openpyxl.load_workbook(file_object)
                sheet = workbook.active
                message = '<p>'
                for i in range(2, sheet.max_row + 1):
                    row_empty = False
                    for j in range(1, 5):
                        if not sheet.cell(row=i, column=j).value:
                            row_empty = True
                            break
                    if row_empty:
                        continue
                    try:
                        mod_name = str(sheet.cell(row=i, column=1).value).strip()
                        date = sheet.cell(row=i, column=2).value.date()
                        user = str(sheet.cell(row=i, column=3).value).strip()
                        rule_str = str(sheet.cell(row=i, column=4).value).strip()
                        action_str = str(sheet.cell(row=i, column=5).value).strip()
                        notes = str(sheet.cell(row=i, column=6).value).strip()

                        mod, mod_created = User.objects.get_or_create(username=mod_name)
                        if mod_created:
                            message += f'[{i}] Mod `{mod}` created.<br />'
                        rule, rule_created = Rule.objects.get_or_create(name=rule_str)
                        if rule_created:
                            message += f'[{i}] Rule `{rule}` created.<br />'
                        action, length = process_action(action_str)

                        Entry.objects.create(
                            moderator=mod,
                            date=date,
                            user=user,
                            rule=rule,
                            action=action,
                            ban_length=length,
                            notes=notes,
                        )

                    except AttributeError as ex:
                        message += f'There was an error processing line {i}.<br />'
                        logger.error(f'On line {i}: {ex}')
                        success = False
                        continue

                entry_count = Entry.objects.count()
                message += f'</p><p>Imported {entry_count} entries.</p>'

        return render(request, 'entries/complete.html',
                      {'success': success, 'message': format_html(message), 'filename': file_object.name})


@login_required
@require_http_methods(['POST'])
def ban_check(request):
    user = request.POST['user']
    if len(user) == 0:
        count = 0
    else:
        if user[:2] == 'u/':
            user = user[2:]
        elif user[:3] == '/u/':
            user = user[3:]
        count = Entry.objects.filter(user__iexact=user).count()
    return render(request, 'entries/_found.html', {'user': user, 'count': count})


@require_http_methods(['GET'])
def health_check(request):
    return HttpResponse('OK', content_type='text/plain', status=200)


class StatsView(TemplateView, LoginRequiredMixin):
    template_name = 'entries/stats.html'


@login_required
def get_mod_actions_chart(request, days: int = 30):
    end_date = now()
    start_date = end_date - timedelta(days=days)

    entries = Entry.objects.filter(date__range=[start_date, end_date])
    active_mods = User.objects.filter(is_active=True, )

    return JsonResponse({
        'title': f'Actions by Mod for the past {days} days',
        'data': {
            'labels': [mod.username for mod in active_mods],
            'datasets': [{
                'label': 'Actions Logged',
                'backgroundColor': [COLOR_PALETTE[i] if i < len(COLOR_PALETTE) else
                                    COLOR_PALETTE[i % len(COLOR_PALETTE)] for i in range(len(active_mods))],
                'data': [entries.filter(moderator__username=mod.username).count() for mod in active_mods]
            }]
        }
    })


@login_required
def get_rules_chart(request, days: int = 30):
    end_date = now()
    start_date = end_date - timedelta(days=days)

    entries = Entry.objects.filter(date__range=[start_date, end_date])
    rules = Rule.objects.all()

    return JsonResponse({
        'title': f'Actions by Rule for the Past {days} day',
        'data': {
            'labels': [rule.name for rule in rules],
            'datasets': [{
                'label': 'Actions Logged',
                'backgroundColor': COLOR_PALETTE[1],
                'data': [entries.filter(rule=rule).count() for rule in rules]
            }]
        }
    })
