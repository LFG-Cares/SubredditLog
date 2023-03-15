import openpyxl
from django.core.management.base import BaseCommand, CommandError

from accounts.models import User
from entries.helpers import process_action
from entries.models import Entry, Rule

COLUMN_MAP = {
    1: 'Mod',
    2: 'Date',
    3: 'User',
    4: 'Rule',
    5: 'Action',
    6: 'Notes',
}


class Command(BaseCommand):
    help = 'Imports an Excel file containing an existing modlog into the database.'

    def add_arguments(self, parser):
        parser.add_argument('filename', nargs=1, type=str, help='The XLSX file to load')
        parser.add_argument('--verbose', action='store_true', help='Provide verbose output')

    def handle(self, *args, **options):
        filename = options['filename'][0]
        verbose = options['verbose']

        if Entry.objects.count() > 0:
            self.stderr.write(
                self.style.ERROR('The database has existing log entries. Aborting.')
            )
            return

        try:
            workbook = openpyxl.load_workbook(filename)
            sheet = workbook.active

            for i in range(2, sheet.max_row+1):
                row_empty = False
                for j in range(1, 5):
                    if not sheet.cell(row=i, column=j).value:
                        if verbose:
                            self.stdout.write(self.style.NOTICE(f'[{i}] Skipping row due to blank {COLUMN_MAP[j]}.'))
                        row_empty = True
                        break
                if row_empty:
                    continue
                try:
                    mod_name = str(sheet.cell(row=i, column=1).value).strip()
                    date = sheet.cell(row=i, column=2).value.date()
                    user = str(sheet.cell(row=i, column=3).value).strip()
                    rule_str = str(sheet.cell(row=i, column=4).value).strip()
                    action_str = str(sheet.cell(row=i, column=5).value).strip()
                    notes = str(sheet.cell(row=i, column=6).value).strip()

                    mod, mod_created = User.objects.get_or_create(username=mod_name)
                    if mod_created:
                        self.stdout.write(self.style.WARNING(f'[{i}] Mod `{mod}` created.'))
                    rule, rule_created = Rule.objects.get_or_create(name=rule_str)
                    if rule_created:
                        self.stdout.write(self.style.WARNING(f'[{i}] Rule `{rule}` created.'))
                    action, length = process_action(action_str)

                    entry = Entry.objects.create(
                        moderator=mod,
                        date=date,
                        user=user,
                        rule=rule,
                        action=action,
                        ban_length=length,
                        notes=notes,
                    )

                    if verbose:
                        self.stdout.write(self.style.NOTICE(f'[{i}] Created {entry}.'))

                except CommandError as ex:
                    self.stderr.write(
                        self.style.ERROR(f'Unable to process action on row {i}: {ex}')
                    )
                    continue
                except AttributeError as ex:
                    self.stderr.write(
                        self.style.ERROR(f'Encountered error processing row {i}: {ex}')
                    )
                    continue
                except Exception as ex:
                    self.stderr.write(
                        self.style.ERROR(f'Encountered error processing row {i}: {ex}')
                    )
                    continue

        except OSError as ex:
            raise CommandError(ex)

        entry_count = Entry.objects.count()
        self.stdout.write(self.style.SUCCESS(f'Loaded {entry_count} log entries.'))
